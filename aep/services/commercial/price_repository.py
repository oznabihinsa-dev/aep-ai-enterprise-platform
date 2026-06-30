from pathlib import Path
from openpyxl import load_workbook


class PriceRepository:
    def __init__(self, file_path: str | Path, year: str = "2026"):
        self.file_path = Path(file_path)
        self.year = year

    def find_price(self, destination: str, coal_mark: str, vat: str = "без НДС") -> dict:
        workbook = load_workbook(self.file_path, data_only=True)

        if self.year not in workbook.sheetnames:
            raise ValueError(f"Лист {self.year} не найден в калькуляции")

        sheet = workbook[self.year]

        coal_mark = coal_mark.lower().replace(" ", "")
        vat = vat.lower().strip()
        destination = destination.lower().strip()

        target_column = None

        for col in range(2, sheet.max_column + 1):
            mark = sheet.cell(row=4, column=col).value
            vat_type = sheet.cell(row=6, column=col).value

            if mark is None or vat_type is None:
                continue

            mark_normalized = str(mark).lower().replace(" ", "")
            vat_normalized = str(vat_type).lower().strip()

            if mark_normalized == coal_mark and vat_normalized == vat:
                target_column = col
                break

        if target_column is None:
            raise ValueError(f"Не найден столбец для марки {coal_mark} и типа цены {vat}")

        for row in range(7, sheet.max_row + 1):
            station = sheet.cell(row=row, column=1).value

            if station is None:
                continue

            if str(station).lower().strip() == destination:
                price = sheet.cell(row=row, column=target_column).value

                if price is None or price == "-":
                    raise ValueError(f"Цена для {destination} / {coal_mark} отсутствует")

                return {
                    "year": self.year,
                    "destination": station,
                    "coal_mark": sheet.cell(row=4, column=target_column).value,
                    "fraction": sheet.cell(row=5, column=target_column).value,
                    "vat_type": sheet.cell(row=6, column=target_column).value,
                    "price": float(price),
                    "column": target_column,
                    "row": row,
                }

        raise ValueError(f"Не найдена станция '{destination}'")