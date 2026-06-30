from openpyxl import load_workbook

wb = load_workbook(
    "sample_data/price/Расчет цены на 2026 год.xlsx",
    data_only=True
)

print("Листы книги:")
print(wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    print("=" * 60)
    print(f"Лист: {sheet_name}")
    print("=" * 60)

    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        print(row)

    print()